#!/usr/bin/env python
#
import openpyxl as px

wb = px.load_workbook('DATA/presidents.xlsx')
ws = wb.get_sheet_by_name('US Presidents')

for i in range(1, ws.max_column + 1):
    print(ws.cell(row=1, column=i).value)
